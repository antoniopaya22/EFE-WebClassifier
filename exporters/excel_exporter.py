from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import openpyxl


def export_excel(matrix, url_keywords, url_categories):
    file = openpyxl.load_workbook('templates/Inventario_Template.xlsx')
    # Inventario
    inventario = file['Inventario y auditoria']
    col_tema = 1
    col_subtema = 2
    col_keywords = 3
    col_url = 4
    row = 2
    for k in url_keywords:
        inventario.cell(row=row, column=col_keywords).value = ','.join(url_keywords[k])
        inventario.cell(row=row, column=col_url).value = k
        inventario.cell(row=row, column=col_keywords).alignment = Alignment(wrapText=True)
        inventario.cell(row=row, column=col_url).alignment = Alignment(wrapText=True)
        row = row + 1
    row = 2
    for k in url_categories:
        inventario.cell(row=row, column=col_tema).value = url_categories[k][0]
        inventario.cell(row=row, column=col_subtema).value = url_categories[k][1]
        inventario.cell(row=row, column=col_tema).alignment = Alignment(wrapText=True)
        inventario.cell(row=row, column=col_subtema).alignment = Alignment(wrapText=True)
        row = row + 1
    matriz_sheet = file['Matriz de similitud']
    for i in range(len(matrix)):
        for j in range(len(matrix[i])):
            matriz_sheet.cell(row=i+1, column=j+1).value = matrix[i][j]
            matriz_sheet.cell(row=i+1, column=j+1).alignment = Alignment(wrapText=True)
            if j == 0 or i == 0:
                matriz_sheet.cell(row=i + 1, column=j + 1).fill = PatternFill("solid", fgColor="4DE83C")
            elif matrix[i][j] == 1:
                matriz_sheet.cell(row=i + 1, column=j + 1).fill = PatternFill("solid", fgColor="DDDDDD")
            elif matrix[i][j] > 0.7:
                matriz_sheet.cell(row=i + 1, column=j + 1).fill = PatternFill("solid", fgColor="E86B3C")
            else:
                matriz_sheet.cell(row=i + 1, column=j + 1).fill = PatternFill("solid", fgColor="FFFFFF")
    file.save('out/export.xlsx')
